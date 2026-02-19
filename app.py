import os
import re
import time
import secrets
from datetime import datetime
from collections import defaultdict
from flask import Flask, render_template, request, jsonify, redirect, session, Response, send_file
import pytz
import psycopg2
from psycopg2.extras import RealDictCursor
from psycopg2 import pool

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "CAMBIA_ESTA_CLAVE_EN_PRODUCCION")

# ===============================
# CONFIGURACIÓN DE SESIÓN
# ===============================
app.config['PERMANENT_SESSION_LIFETIME'] = 8 * 60 * 60  # 8 horas en segundos
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = True  # Cookie solo viaja por HTTPS

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
TIMEZONE = pytz.timezone('America/Lima')
DATABASE_URL = os.environ.get("DATABASE_URL")
print("💾 Usando PostgreSQL (Azure)")

# ===============================
# CONNECTION POOL
# ===============================
db_pool = pool.ThreadedConnectionPool(
    minconn=2,
    maxconn=10,
    dsn=DATABASE_URL,
    cursor_factory=RealDictCursor
)

def get_conn():
    return db_pool.getconn()

def release_conn(conn):
    db_pool.putconn(conn)

class PooledConn:
    def __init__(self):
        self.conn = None
    def __enter__(self):
        self.conn = get_conn()
        return self.conn
    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type:
            try:
                self.conn.rollback()
            except Exception:
                pass
        release_conn(self.conn)

# ===============================
# RATE LIMITING — LOGIN
# ===============================
_login_attempts = defaultdict(list)
LOGIN_MAX_INTENTOS = 5
LOGIN_VENTANA_SEG = 5 * 60

def _limpiar_intentos_viejos(ip):
    ahora = time.time()
    _login_attempts[ip] = [t for t in _login_attempts[ip] if ahora - t < LOGIN_VENTANA_SEG]

def esta_bloqueado(ip):
    _limpiar_intentos_viejos(ip)
    return len(_login_attempts[ip]) >= LOGIN_MAX_INTENTOS

def registrar_intento(ip):
    _login_attempts[ip].append(time.time())

def segundos_restantes(ip):
    if not _login_attempts[ip]:
        return 0
    mas_viejo = _login_attempts[ip][0]
    return max(0, int(LOGIN_VENTANA_SEG - (time.time() - mas_viejo)))

# ===============================
# CSRF — token por sesión
# ===============================
def generar_csrf_token():
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
    return session['csrf_token']

def validar_csrf(request):
    token_sesion = session.get('csrf_token')
    if not token_sesion:
        return False
    token_req = (
        request.headers.get('X-CSRF-Token')
        or request.form.get('csrf_token')
        or (request.get_json(silent=True) or {}).get('csrf_token')
    )
    return token_req == token_sesion

# ===============================
# VERIFICACIÓN DE SESIÓN
# ===============================
def sesion_activa(rol_requerido=None):
    if 'usuario' not in session:
        return False
    if rol_requerido and session.get('rol') != rol_requerido:
        return False
    login_time = session.get('login_time', 0)
    if time.time() - login_time > 8 * 3600:
        session.clear()
        return False
    return True


def now_peru():
    return datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")


def registrar_log(usuario, accion):
    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                  INSERT INTO logs (fecha, usuario, accion)
                  VALUES (%s, %s, %s)
                """, (now_peru(), usuario, accion))
            conn.commit()
    except Exception as e:
        print(f"Error al registrar log: {e}")


def ensure_column(conn, table, column, ddl):
    with conn.cursor() as cur:
        cur.execute("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = %s AND column_name = %s
        """, (table, column))
        exists = cur.fetchone()
        if not exists:
            cur.execute(ddl)
            conn.commit()


def init_db():
    with PooledConn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
            CREATE TABLE IF NOT EXISTS postulantes (
              id SERIAL PRIMARY KEY,
              created_at TEXT NOT NULL,
              convocatoria TEXT NOT NULL,
              apellidos TEXT NOT NULL,
              nombres TEXT NOT NULL,
              tipo_documento TEXT NOT NULL,
              numero_documento TEXT NOT NULL,
              fecha_nacimiento TEXT NOT NULL,
              sexo TEXT NOT NULL,
              celular TEXT NOT NULL,
              correo TEXT NOT NULL
            );
            """)
            conn.commit()

        ensure_column(conn, "postulantes", "validado",
            "ALTER TABLE postulantes ADD COLUMN validado INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "postulantes", "fuerzas_armadas",
            "ALTER TABLE postulantes ADD COLUMN fuerzas_armadas TEXT")
        ensure_column(conn, "postulantes", "tiene_discapacidad",
            "ALTER TABLE postulantes ADD COLUMN tiene_discapacidad TEXT")
        ensure_column(conn, "postulantes", "tipo_discapacidad",
            "ALTER TABLE postulantes ADD COLUMN tipo_discapacidad TEXT")
        ensure_column(conn, "postulantes", "area",
            "ALTER TABLE postulantes ADD COLUMN area TEXT")
        ensure_column(conn, "postulantes", "usuario_atendio",
            "ALTER TABLE postulantes ADD COLUMN usuario_atendio TEXT")
        ensure_column(conn, "postulantes", "fecha_atencion",
            "ALTER TABLE postulantes ADD COLUMN fecha_atencion TEXT")

        with conn.cursor() as cur:
            cur.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
              id SERIAL PRIMARY KEY,
              username TEXT UNIQUE NOT NULL,
              password TEXT NOT NULL,
              rol TEXT NOT NULL,
              activo INTEGER NOT NULL DEFAULT 1,
              created_at TEXT NOT NULL
            );
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS logs (
              id SERIAL PRIMARY KEY,
              fecha TEXT NOT NULL,
              usuario TEXT NOT NULL,
              accion TEXT NOT NULL
            );
            """)
            conn.commit()

            cur.execute("SELECT id FROM usuarios WHERE username='admin'")
            if not cur.fetchone():
                cur.execute("""
                  INSERT INTO usuarios (username, password, rol, activo, created_at)
                  VALUES ('admin', %s, 'admin', 1, %s)
                """, (os.environ.get("ADMIN_PASSWORD", "Admin2026@Muni!"), now_peru()))
                conn.commit()

            cur.execute("""
                CREATE TABLE IF NOT EXISTS configuracion (
                  clave TEXT PRIMARY KEY,
                  valor TEXT NOT NULL
                );
            """)
            cur.execute("""
                INSERT INTO configuracion (clave, valor)
                VALUES ('convocatoria_activa', 'true')
                ON CONFLICT (clave) DO NOTHING;
            """)

            cur.execute("""
                CREATE TABLE IF NOT EXISTS sesiones_activas (
                  username TEXT PRIMARY KEY,
                  ultimo_latido TEXT NOT NULL
                );
            """)
            conn.commit()


def crear_indices():
    with PooledConn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE UNIQUE INDEX IF NOT EXISTS idx_documento_unico
                ON postulantes(numero_documento, tipo_documento)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_usuario_atendio
                ON postulantes(usuario_atendio)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_sexo
                ON postulantes(sexo)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_convocatoria
                ON postulantes(convocatoria)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_area
                ON postulantes(area)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_usuario_sexo
                ON postulantes(usuario_atendio, sexo)
            """)
            conn.commit()


init_db()
crear_indices()


# ===============================
# WEB
# ===============================
@app.get("/")
def home():
    return render_template("Formulario.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if sesion_activa("admin"):
        return redirect("/admin")
    if sesion_activa("usuario"):
        return redirect("/usuario")

    csrf_token = generar_csrf_token()

    if request.method == "POST":
        ip = request.remote_addr

        if not validar_csrf(request):
            return render_template("Login.html", error="Solicitud inválida. Recarga la página.", csrf_token=csrf_token)

        if esta_bloqueado(ip):
            seg = segundos_restantes(ip)
            minutos = seg // 60
            segundos = seg % 60
            return render_template("Login.html",
                error=f"Demasiados intentos fallidos. Intenta en {minutos}m {segundos}s.",
                csrf_token=csrf_token)

        usuario = (request.form.get("usuario") or "").strip()
        password = (request.form.get("password") or "").strip()

        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                  SELECT username, rol FROM usuarios
                  WHERE username=%s AND password=%s AND activo=1
                """, (usuario, password))
                row = cur.fetchone()

        if row:
            _login_attempts[ip].clear()
            session.clear()
            session.permanent = True
            session["usuario"] = row["username"]
            session["rol"] = row["rol"]
            session["admin"] = row["rol"] == "admin"
            session["login_time"] = time.time()
            session['csrf_token'] = secrets.token_hex(32)
            rol = row["rol"]
            username = row["username"]
            registrar_log(username, "Inicio de sesión")
            return redirect("/admin" if rol == "admin" else "/usuario")

        registrar_intento(ip)
        intentos_restantes = LOGIN_MAX_INTENTOS - len(_login_attempts[ip])
        if intentos_restantes > 0:
            msg = f"Credenciales incorrectas. Te quedan {intentos_restantes} intento(s)."
        else:
            seg = segundos_restantes(ip)
            msg = f"Demasiados intentos. Bloqueado por {seg // 60}m {seg % 60}s."

        return render_template("Login.html", error=msg, csrf_token=csrf_token)

    return render_template("Login.html", csrf_token=csrf_token)


@app.get("/logout")
def logout():
    usuario = session.get("usuario", "Desconocido")
    # Eliminar de sesiones activas inmediatamente al cerrar sesión
    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM sesiones_activas WHERE username = %s", (usuario,))
                conn.commit()
    except Exception:
        pass
    registrar_log(usuario, "Cerró sesión")
    session.clear()
    return redirect("/login")


@app.get("/admin")
def admin():
    if not sesion_activa("admin"):
        return redirect("/login")

    with PooledConn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT username, rol, activo AS estado FROM usuarios")
            usuarios = cur.fetchall()
            cur.execute("SELECT fecha, usuario, accion FROM logs ORDER BY id DESC LIMIT 25")
            logs = cur.fetchall()
            cur.execute("SELECT COUNT(*) as total FROM logs")
            total_logs = cur.fetchone()["total"]

    return render_template(
        "admin.html",
        usuarios=usuarios,
        usuario=session.get("usuario", "Admin"),
        logs=logs,
        total_logs=total_logs,
        csrf_token=generar_csrf_token()
    )


@app.get("/usuario")
def usuario_panel():
    if not sesion_activa("usuario"):
        return redirect("/login")

    return render_template(
        "usuario.html",
        usuario=session.get("usuario"),
        csrf_token=generar_csrf_token()
    )


# ===============================
# API — helper de autorización
# ===============================
def require_rol(*roles):
    if 'usuario' not in session:
        return jsonify({"ok": False, "error": "No autorizado"}), 403
    if session.get('rol') not in roles:
        return jsonify({"ok": False, "error": "No autorizado"}), 403
    login_time = session.get('login_time', 0)
    if time.time() - login_time > 8 * 3600:
        session.clear()
        return jsonify({"ok": False, "error": "Sesión expirada"}), 401
    return None

def require_csrf():
    if not validar_csrf(request):
        return jsonify({"ok": False, "error": "Token CSRF inválido"}), 403
    return None


# ===============================
# API
# ===============================
@app.get("/api/health")
def health():
    return jsonify({"ok": True, "db": "postgresql"})


@app.post("/api/verificar-postulante")
def verificar_postulante():
    data = request.get_json(silent=True) or {}
    numero_documento = (data.get("numero_documento") or "").strip()
    tipo_documento = (data.get("tipo_documento") or "").strip()

    if not numero_documento or not tipo_documento:
        return jsonify({"ok": False, "error": "Datos incompletos"}), 400

    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT convocatoria, area, apellidos, nombres, created_at
                    FROM postulantes
                    WHERE numero_documento = %s AND tipo_documento = %s
                    LIMIT 1
                """, (numero_documento, tipo_documento))
                resultado = cur.fetchone()

                if resultado:
                    return jsonify({
                        "ok": True,
                        "existe": True,
                        "convocatoria": resultado["convocatoria"],
                        "area": resultado["area"],
                        "apellidos": resultado["apellidos"],
                        "nombres": resultado["nombres"],
                        "fecha_registro": resultado["created_at"]
                    }), 200
                else:
                    return jsonify({"ok": True, "existe": False}), 200

    except Exception as e:
        print(f"❌ Error al verificar postulante: {str(e)}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post("/api/submit")
def submit():
    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT valor FROM configuracion WHERE clave = 'convocatoria_activa'")
                row = cur.fetchone()
                if row and row['valor'] == 'false':
                    return jsonify({"ok": False, "cerrado": True,
                                    "error": "La convocatoria ha finalizado. Ya no se aceptan registros."}), 403
    except Exception as e:
        print(f"Error verificando estado convocatoria: {e}")

    data = request.get_json(silent=True) or {}

    area = (data.get("area") or "").strip()
    convocatoria = (data.get("convocatoria") or "").strip()
    apellidos = (data.get("apellidos") or "").strip().upper()
    nombres = (data.get("nombres") or "").strip().upper()
    tipo_documento = (data.get("tipo_documento") or "").strip()
    numero_documento = (data.get("numero_documento") or "").strip()
    fecha_nacimiento = (data.get("fecha_nacimiento") or "").strip()
    sexo = (data.get("sexo") or "").strip()
    celular = (data.get("celular") or "").strip()
    correo = (data.get("correo") or "").strip()
    fuerzas_armadas = (data.get("fuerzas_armadas") or "").strip()
    tiene_discapacidad = (data.get("tiene_discapacidad") or "").strip()
    tipo_discapacidad = (data.get("tipo_discapacidad") or "").strip()

    if not all([area, convocatoria, apellidos, nombres, tipo_documento,
                numero_documento, fecha_nacimiento,
                sexo, celular, correo, fuerzas_armadas, tiene_discapacidad]):
        return jsonify({"ok": False, "error": "Completa todos los campos"}), 400

    if not EMAIL_RE.match(correo):
        return jsonify({"ok": False, "error": "Correo inválido"}), 400

    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT convocatoria, area FROM postulantes
                    WHERE numero_documento = %s AND tipo_documento = %s
                    LIMIT 1
                """, (numero_documento, tipo_documento))
                existe = cur.fetchone()

                if existe:
                    return jsonify({
                        "ok": False,
                        "error": f"El {tipo_documento} {numero_documento} ya está registrado en: {existe['convocatoria']}"
                    }), 400

                cur.execute("""
                  INSERT INTO postulantes
                  (created_at, area, convocatoria, apellidos, nombres, tipo_documento,
                   numero_documento, fecha_nacimiento, sexo, celular, correo,
                   fuerzas_armadas, tiene_discapacidad, tipo_discapacidad, validado)
                  VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0)
                """, (
                    now_peru(), area, convocatoria, apellidos, nombres, tipo_documento,
                    numero_documento, fecha_nacimiento, sexo, celular, correo,
                    fuerzas_armadas, tiene_discapacidad, tipo_discapacidad
                ))
                conn.commit()
                print(f"✅ Postulante registrado: {apellidos}, {nombres}")

        return jsonify({"ok": True})

    except Exception as e:
        print(f"❌ Error al registrar: {str(e)}")
        return jsonify({"ok": False, "error": str(e)}), 500


# -----------------------------------------------
# POLLING
# -----------------------------------------------
@app.get("/api/postulantes/pendientes-nuevos")
def postulantes_pendientes_nuevos():
    err = require_rol("admin", "usuario")
    if err: return err

    after_id = request.args.get("after_id", 0, type=int)

    with PooledConn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, area, convocatoria, apellidos, nombres, tipo_documento,
                       numero_documento, fecha_nacimiento, sexo, celular, correo,
                       fuerzas_armadas, tiene_discapacidad, tipo_discapacidad, created_at
                FROM postulantes
                WHERE id > %s AND usuario_atendio IS NULL
                ORDER BY id ASC
            """, (after_id,))
            rows = cur.fetchall()

    items = [dict(r) for r in rows]
    return jsonify({"ok": True, "items": items})


@app.post("/api/postulantes/datos-atendidos")
def datos_atendidos():
    err = require_rol("admin", "usuario")
    if err: return err

    data = request.get_json(silent=True) or {}
    ids = data.get("ids", [])

    if not ids:
        return jsonify({"ok": True, "items": []})

    try:
        ids = [int(i) for i in ids]
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "IDs inválidos"}), 400

    placeholders = ','.join(['%s'] * len(ids))

    with PooledConn() as conn:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT id, area, convocatoria, apellidos, nombres, tipo_documento,
                       numero_documento, fecha_nacimiento, sexo, celular, correo,
                       fuerzas_armadas, tiene_discapacidad, tipo_discapacidad,
                       created_at, usuario_atendio, fecha_atencion
                FROM postulantes
                WHERE id IN ({placeholders}) AND usuario_atendio IS NOT NULL
            """, tuple(ids))
            rows = cur.fetchall()

    return jsonify({"ok": True, "items": [dict(r) for r in rows]})


@app.get("/api/postulantes/atendidos-ids")
def postulantes_atendidos_ids():
    err = require_rol("admin", "usuario")
    if err: return err

    after_id = request.args.get("after_id", 0, type=int)

    with PooledConn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, usuario_atendio
                FROM postulantes
                WHERE id <= %s AND usuario_atendio IS NOT NULL
            """, (after_id,))
            rows = cur.fetchall()

    items = [{"id": r["id"], "usuario_atendio": r["usuario_atendio"]} for r in rows]
    return jsonify({"ok": True, "items": items})


@app.get("/api/postulantes/atendidos-nuevos")
def postulantes_atendidos_nuevos():
    err = require_rol("admin")
    if err: return err

    after_id = request.args.get("after_id", 0, int)

    with PooledConn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, area, convocatoria, apellidos, nombres, tipo_documento,
                       numero_documento, fecha_nacimiento, sexo, celular, correo,
                       fuerzas_armadas, tiene_discapacidad, tipo_discapacidad,
                       created_at, usuario_atendio, fecha_atencion
                FROM postulantes
                WHERE id > %s AND usuario_atendio IS NOT NULL
                ORDER BY id ASC
            """, (after_id,))
            rows = cur.fetchall()

    items = [dict(r) for r in rows]
    return jsonify({"ok": True, "items": items})


@app.get("/api/postulantes/registrados")
def postulantes_registrados():
    err = require_rol("admin", "usuario")
    if err: return err

    with PooledConn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, area, convocatoria, apellidos, nombres, tipo_documento,
                       numero_documento, fecha_nacimiento, sexo, celular, correo,
                       fuerzas_armadas, tiene_discapacidad, tipo_discapacidad, created_at
                FROM postulantes
                WHERE usuario_atendio IS NULL
                ORDER BY created_at DESC
            """)
            rows = cur.fetchall()

    items = [dict(r) for r in rows]
    return jsonify({"ok": True, "items": items})


@app.get("/api/estadisticas")
def estadisticas():
    err = require_rol("admin")
    if err: return err

    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT
                        COUNT(*) FILTER (WHERE usuario_atendio IS NULL AND sexo = 'Femenino')  AS reg_mujeres,
                        COUNT(*) FILTER (WHERE usuario_atendio IS NULL AND sexo = 'Masculino') AS reg_hombres,
                        COUNT(*) FILTER (WHERE usuario_atendio IS NOT NULL AND sexo = 'Femenino')  AS rec_mujeres,
                        COUNT(*) FILTER (WHERE usuario_atendio IS NOT NULL AND sexo = 'Masculino') AS rec_hombres
                    FROM postulantes
                """)
                totales = cur.fetchone()

                cur.execute("""
                    SELECT area, COUNT(*) as total FROM postulantes
                    WHERE area IS NOT NULL AND area != ''
                    GROUP BY area ORDER BY total DESC
                """)
                rows_area = cur.fetchall()
                por_area = {row["area"]: row["total"] for row in rows_area}

        return jsonify({
            "ok": True,
            "registrados_mujeres": int(totales["reg_mujeres"]),
            "registrados_hombres": int(totales["reg_hombres"]),
            "recibidos_mujeres": int(totales["rec_mujeres"]),
            "recibidos_hombres": int(totales["rec_hombres"]),
            "por_area": por_area
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post("/api/eliminar/<int:pid>")
def api_eliminar(pid):
    err = require_rol("admin", "usuario")
    if err: return err
    err2 = require_csrf()
    if err2: return err2

    with PooledConn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT apellidos, nombres FROM postulantes WHERE id = %s", (pid,))
            p = cur.fetchone()
            cur.execute("DELETE FROM postulantes WHERE id=%s", (pid,))
            conn.commit()

    if p:
        registrar_log(session.get("usuario", "?"), f"Eliminó a {p['apellidos']}, {p['nombres']} (ID: {pid})")

    return jsonify({"ok": True})


@app.post("/api/editar-postulante")
def editar_postulante():
    err = require_rol("usuario")
    if err: return err
    err2 = require_csrf()
    if err2: return err2

    data = request.get_json(silent=True) or {}
    postulante_id = data.get("id")
    if not postulante_id:
        return jsonify({"ok": False, "error": "ID no proporcionado"}), 400

    area = (data.get("area") or "").strip() or None
    convocatoria = (data.get("convocatoria") or "").strip()
    apellidos = (data.get("apellidos") or "").strip().upper()
    nombres = (data.get("nombres") or "").strip().upper()
    tipo_documento = (data.get("tipo_documento") or "").strip()
    numero_documento = (data.get("numero_documento") or "").strip()
    fecha_nacimiento = (data.get("fecha_nacimiento") or "").strip()
    sexo = (data.get("sexo") or "").strip()
    celular = (data.get("celular") or "").strip()
    correo = (data.get("correo") or "").strip()
    fuerzas_armadas = (data.get("fuerzas_armadas") or "").strip() or None
    tiene_discapacidad = (data.get("tiene_discapacidad") or "").strip() or None
    tipo_discapacidad = (data.get("tipo_discapacidad") or "").strip() or None

    if not all([convocatoria, apellidos, nombres, tipo_documento,
                numero_documento, fecha_nacimiento, sexo, celular, correo]):
        return jsonify({"ok": False, "error": "Completa todos los campos obligatorios"}), 400

    if not EMAIL_RE.match(correo):
        return jsonify({"ok": False, "error": "Correo inválido"}), 400

    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT usuario_atendio FROM postulantes WHERE id = %s", (postulante_id,))
                postulante = cur.fetchone()

                if not postulante:
                    return jsonify({"ok": False, "error": "Postulante no encontrado"}), 404

                if postulante["usuario_atendio"] is not None:
                    return jsonify({
                        "ok": False,
                        "error": "No se puede editar un postulante que ya fue atendido"
                    }), 400

                cur.execute("""
                    SELECT id FROM postulantes
                    WHERE numero_documento = %s AND tipo_documento = %s AND id != %s
                    LIMIT 1
                """, (numero_documento, tipo_documento, postulante_id))

                if cur.fetchone():
                    return jsonify({
                        "ok": False,
                        "error": f"El {tipo_documento} {numero_documento} ya está registrado en otro postulante"
                    }), 400

                cur.execute("""
                    UPDATE postulantes
                    SET area = %s, convocatoria = %s, apellidos = %s, nombres = %s,
                        tipo_documento = %s, numero_documento = %s, fecha_nacimiento = %s,
                        sexo = %s, celular = %s, correo = %s, fuerzas_armadas = %s,
                        tiene_discapacidad = %s, tipo_discapacidad = %s
                    WHERE id = %s AND usuario_atendio IS NULL
                """, (
                    area, convocatoria, apellidos, nombres, tipo_documento,
                    numero_documento, fecha_nacimiento, sexo, celular, correo,
                    fuerzas_armadas, tiene_discapacidad, tipo_discapacidad,
                    postulante_id
                ))
                conn.commit()

                if cur.rowcount == 0:
                    return jsonify({
                        "ok": False,
                        "error": "No se pudo actualizar. El postulante puede haber sido atendido."
                    }), 400

                registrar_log(session.get("usuario"), f"Editó datos de {apellidos}, {nombres} (ID: {postulante_id})")
                return jsonify({"ok": True})

    except Exception as e:
        print(f"❌ Error al editar postulante: {str(e)}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post("/api/recibir-postulante")
def recibir_postulante():
    err = require_rol("usuario")
    if err: return err
    err2 = require_csrf()
    if err2: return err2

    data = request.get_json()
    postulante_id = data.get("id")

    if not postulante_id:
        return jsonify({"ok": False, "error": "ID no proporcionado"}), 400

    usuario_actual = session.get("usuario")
    fecha_actual = now_peru()

    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT apellidos, nombres FROM postulantes WHERE id = %s", (postulante_id,))
                postulante = cur.fetchone()

                if not postulante:
                    return jsonify({"ok": False, "error": "Postulante no encontrado"}), 404

                cur.execute("""
                    UPDATE postulantes
                    SET usuario_atendio = %s, fecha_atencion = %s
                    WHERE id = %s AND usuario_atendio IS NULL
                """, (usuario_actual, fecha_actual, postulante_id))
                conn.commit()

                if cur.rowcount == 0:
                    cur.execute("SELECT usuario_atendio FROM postulantes WHERE id = %s", (postulante_id,))
                    row = cur.fetchone()
                    quien = row["usuario_atendio"] if row else "otro usuario"
                    return jsonify({
                        "ok": False,
                        "ya_tomado": True,
                        "error": f"Ya fue recibido por {quien}"
                    }), 409

        registrar_log(usuario_actual, f"Recibió a {postulante['apellidos']}, {postulante['nombres']}")
        return jsonify({"ok": True})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ===============================
# USUARIOS
# ===============================
@app.post("/api/activar-usuario")
def activar_usuario():
    err = require_rol("admin")
    if err: return err
    err2 = require_csrf()
    if err2: return err2

    data = request.get_json()
    username = data.get("username")
    with PooledConn() as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE usuarios SET activo = 1 WHERE username = %s", (username,))
            conn.commit()
    return jsonify({"ok": True})


@app.post("/api/desactivar-usuario")
def desactivar_usuario():
    err = require_rol("admin")
    if err: return err
    err2 = require_csrf()
    if err2: return err2

    data = request.get_json()
    username = data.get("username")
    if username == "admin":
        return jsonify({"ok": False, "error": "No se puede desactivar admin"}), 400
    with PooledConn() as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE usuarios SET activo = 0 WHERE username = %s", (username,))
            conn.commit()
    return jsonify({"ok": True})


@app.post("/api/crear-usuario")
def crear_usuario():
    err = require_rol("admin")
    if err: return err
    err2 = require_csrf()
    if err2: return err2

    data = request.get_json()
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    rol = data.get("rol", "usuario")
    if not username or not password:
        return jsonify({"ok": False, "error": "Datos incompletos"}), 400
    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                  INSERT INTO usuarios (username, password, rol, activo, created_at)
                  VALUES (%s, %s, %s, 1, %s)
                """, (username, password, rol, now_peru()))
                conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post("/api/eliminar-usuario")
def eliminar_usuario():
    err = require_rol("admin")
    if err: return err
    err2 = require_csrf()
    if err2: return err2

    data = request.get_json()
    username = data.get("username", "").strip()
    if username == "admin":
        return jsonify({"ok": False, "error": "No se puede eliminar el usuario admin"}), 400
    if not username:
        return jsonify({"ok": False, "error": "Usuario no especificado"}), 400
    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM usuarios WHERE username = %s", (username,))
                conn.commit()
                if cur.rowcount == 0:
                    return jsonify({"ok": False, "error": "Usuario no encontrado"}), 404
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post("/api/cambiar-password")
def cambiar_password():
    err = require_rol("admin")
    if err: return err
    err2 = require_csrf()
    if err2: return err2

    data = request.get_json(silent=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()

    if not username or not password:
        return jsonify({"ok": False, "error": "Datos incompletos"}), 400

    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE usuarios SET password = %s WHERE username = %s", (password, username))
                conn.commit()
                if cur.rowcount == 0:
                    return jsonify({"ok": False, "error": "Usuario no encontrado"}), 404
        registrar_log(session.get("usuario"), f"Cambió contraseña de {username}")
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.get("/api/logs")
def api_logs():
    err = require_rol("admin")
    if err: return err

    pagina = request.args.get("pagina", 1, type=int)
    tam = request.args.get("tam", 25, type=int)
    buscar = request.args.get("buscar", "").strip().lower()

    offset = (pagina - 1) * tam

    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                if buscar:
                    cur.execute("""
                        SELECT fecha, usuario, accion FROM logs
                        WHERE LOWER(usuario) LIKE %s OR LOWER(accion) LIKE %s
                        ORDER BY id DESC LIMIT %s OFFSET %s
                    """, (f"%{buscar}%", f"%{buscar}%", tam, offset))
                    rows = cur.fetchall()
                    cur.execute("""
                        SELECT COUNT(*) as total FROM logs
                        WHERE LOWER(usuario) LIKE %s OR LOWER(accion) LIKE %s
                    """, (f"%{buscar}%", f"%{buscar}%"))
                else:
                    cur.execute("""
                        SELECT fecha, usuario, accion FROM logs
                        ORDER BY id DESC LIMIT %s OFFSET %s
                    """, (tam, offset))
                    rows = cur.fetchall()
                    cur.execute("SELECT COUNT(*) as total FROM logs")
                total = cur.fetchone()["total"]

        return jsonify({"ok": True, "items": [dict(r) for r in rows], "total": total})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post("/api/limpiar-logs")
def limpiar_logs():
    err = require_rol("admin")
    if err: return err
    err2 = require_csrf()
    if err2: return err2

    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM logs")
                eliminados = cur.rowcount
                conn.commit()
        return jsonify({"ok": True, "eliminados": eliminados})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ===============================
# EXPORTACIONES
# ===============================
@app.get("/admin/export/csv")
def export_csv():
    if not sesion_activa("admin"):
        return redirect("/login")

    import csv
    from io import StringIO

    with PooledConn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
              SELECT id, area, convocatoria, apellidos, nombres, tipo_documento,
                     numero_documento, fecha_nacimiento, sexo, celular, correo,
                     fuerzas_armadas, tiene_discapacidad, tipo_discapacidad,
                     created_at, usuario_atendio, fecha_atencion
              FROM postulantes
              WHERE usuario_atendio IS NOT NULL
              ORDER BY fecha_atencion DESC
            """)
            postulantes = cur.fetchall()

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'ID', 'Área', 'Convocatoria', 'Apellidos', 'Nombres', 'Tipo Doc', 'N° Doc',
        'Fecha Nacimiento', 'Sexo', 'Celular', 'Correo', 'FF.AA.',
        'Discapacidad', 'Tipo Discapacidad', 'Fecha Registro',
        'Usuario Atendió', 'Fecha Atención'
    ])
    for p in postulantes:
        writer.writerow([
            p['id'], p['area'] or '', p['convocatoria'], p['apellidos'], p['nombres'],
            p['tipo_documento'], p['numero_documento'], p['fecha_nacimiento'],
            p['sexo'], p['celular'], p['correo'], p['fuerzas_armadas'],
            p['tiene_discapacidad'], p['tipo_discapacidad'] or '',
            p['created_at'], p['usuario_atendio'], p['fecha_atencion']
        ])
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=postulantes_{datetime.now(TIMEZONE).strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )


COLORES_AREA = {
    'GGRD':  {'fondo': 'FFDCE1', 'letra': '7B0000'},
    'GSCGA': {'fondo': 'D6F5E3', 'letra': '0B4D2B'},
    'GFC':   {'fondo': 'D0E8FF', 'letra': '0A2F6B'},
    'GSC':   {'fondo': 'EDE0FF', 'letra': '3B006B'},
    'GDE':   {'fondo': 'FFF3CC', 'letra': '6B4200'},
}

def aplicar_estilo_excel(ws, postulantes, headers, col_area_idx=1):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="003F8F", end_color="003F8F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 32
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    row_align = Alignment(horizontal="left", vertical="center")
    for row_num, p in enumerate(postulantes, 2):
        area_val = p.get('area') or ''
        cfg = COLORES_AREA.get(area_val, {'fondo': 'FFFFFF', 'letra': '1F2937'})
        row_fill = PatternFill(start_color=cfg['fondo'], end_color=cfg['fondo'], fill_type="solid")
        row_font = Font(color=cfg['letra'], size=9)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = row_fill
            cell.font = row_font
            cell.alignment = row_align
            cell.border = border

    for col_idx, col in enumerate(ws.iter_cols(), 1):
        col_letter = get_column_letter(col_idx)
        max_len = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[col_letter].width = max(8, min(45, max_len + 3))

    total_row = len(postulantes) + 2
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color="003F8F", end_color="003F8F", fill_type="solid")
    ws.cell(row=total_row, column=1).font = Font(bold=True, color="FFFFFF", size=9)
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=2).value = len(postulantes)
    ws.cell(row=total_row, column=2).font = Font(bold=True, size=9)

    areas_count = {}
    for p in postulantes:
        a = p.get('area') or 'Sin área'
        areas_count[a] = areas_count.get(a, 0) + 1

    nota = " | ".join(f"{k}: {v}" for k, v in sorted(areas_count.items()))
    ws.cell(row=total_row, column=3, value=nota).font = Font(italic=True, size=8, color="555555")
    ws.merge_cells(start_row=total_row, start_column=3, end_row=total_row, end_column=len(headers))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(postulantes) + 1}"
    ws.freeze_panes = "A2"


@app.get("/admin/export/excel")
def export_excel():
    if not sesion_activa("admin"):
        return redirect("/login")
    try:
        from openpyxl import Workbook
        from io import BytesIO
    except ImportError:
        return jsonify({"ok": False, "error": "openpyxl no está instalado"}), 500

    with PooledConn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
              SELECT id, area, convocatoria, apellidos, nombres, tipo_documento,
                     numero_documento, fecha_nacimiento, sexo, celular, correo,
                     fuerzas_armadas, tiene_discapacidad, tipo_discapacidad,
                     created_at, usuario_atendio, fecha_atencion
              FROM postulantes WHERE usuario_atendio IS NOT NULL ORDER BY area, fecha_atencion
            """)
            postulantes = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Recibidos"

    headers = [
        'ID', 'Área', 'Convocatoria', 'Apellidos', 'Nombres', 'Tipo Doc', 'N° Doc',
        'Fecha Nacimiento', 'Sexo', 'Celular', 'Correo', 'FF.AA.',
        'Discapacidad', 'Tipo Discapacidad', 'Fecha Registro', 'Usuario Atendió', 'Fecha Atención'
    ]

    for row_num, p in enumerate(postulantes, 2):
        vals = [
            p['id'], p['area'] or '', p['convocatoria'], p['apellidos'], p['nombres'],
            p['tipo_documento'], p['numero_documento'], p['fecha_nacimiento'],
            p['sexo'], p['celular'], p['correo'], p['fuerzas_armadas'] or '',
            p['tiene_discapacidad'] or '', p['tipo_discapacidad'] or '',
            p['created_at'], p['usuario_atendio'], p['fecha_atencion']
        ]
        for col_num, val in enumerate(vals, 1):
            ws.cell(row=row_num, column=col_num, value=val)

    aplicar_estilo_excel(ws, postulantes, headers)

    ws_res = wb.create_sheet("Resumen por Área")
    from openpyxl.styles import Font, PatternFill, Alignment
    ws_res.column_dimensions['A'].width = 35
    ws_res.column_dimensions['B'].width = 12
    ws_res.column_dimensions['C'].width = 12
    ws_res.column_dimensions['D'].width = 12
    res_headers = ['Área', 'Total', 'Hombres', 'Mujeres']
    for i, h in enumerate(res_headers, 1):
        c = ws_res.cell(row=1, column=i, value=h)
        c.fill = PatternFill(start_color="003F8F", end_color="003F8F", fill_type="solid")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center")

    resumen = {}
    for p in postulantes:
        a = p['area'] or 'Sin área'
        if a not in resumen:
            resumen[a] = {'total': 0, 'h': 0, 'm': 0}
        resumen[a]['total'] += 1
        if p['sexo'] == 'Masculino': resumen[a]['h'] += 1
        else: resumen[a]['m'] += 1

    NOMBRES_AREA = {
        'GGRD': 'Gestión del Riesgo de Desastres',
        'GSCGA': 'Servicios a la Ciudad y Gestión Ambiental',
        'GFC': 'Fiscalización y Control',
        'GSC': 'Seguridad Ciudadana',
        'GDE': 'Desarrollo Económico',
    }
    for r, (area_key, datos) in enumerate(sorted(resumen.items()), 2):
        cfg = COLORES_AREA.get(area_key, {'fondo': 'F3F4F6', 'letra': '1F2937'})
        fill = PatternFill(start_color=cfg['fondo'], end_color=cfg['fondo'], fill_type="solid")
        nombre_largo = NOMBRES_AREA.get(area_key, area_key)
        for ci, val in enumerate([nombre_largo, datos['total'], datos['h'], datos['m']], 1):
            c = ws_res.cell(row=r, column=ci, value=val)
            c.fill = fill
            c.font = Font(size=9, bold=(ci == 1))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"recibidos_{datetime.now(TIMEZONE).strftime('%Y%m%d_%H%M%S')}.xlsx")


@app.get("/admin/export/excel-pendientes")
def export_excel_pendientes():
    if not sesion_activa("admin"):
        return redirect("/login")
    try:
        from openpyxl import Workbook
        from io import BytesIO
    except ImportError:
        return jsonify({"ok": False, "error": "openpyxl no está instalado"}), 500

    with PooledConn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
              SELECT id, area, convocatoria, apellidos, nombres, tipo_documento,
                     numero_documento, fecha_nacimiento, sexo, celular, correo,
                     fuerzas_armadas, tiene_discapacidad, tipo_discapacidad, created_at
              FROM postulantes WHERE usuario_atendio IS NULL ORDER BY area, created_at
            """)
            postulantes = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Registrados"

    headers = [
        'ID', 'Área', 'Convocatoria', 'Apellidos', 'Nombres', 'Tipo Doc', 'N° Doc',
        'Fecha Nacimiento', 'Sexo', 'Celular', 'Correo', 'FF.AA.',
        'Discapacidad', 'Tipo Discapacidad', 'Fecha Registro'
    ]

    for row_num, p in enumerate(postulantes, 2):
        vals = [
            p['id'], p['area'] or '', p['convocatoria'], p['apellidos'], p['nombres'],
            p['tipo_documento'], p['numero_documento'], p['fecha_nacimiento'],
            p['sexo'], p['celular'], p['correo'], p['fuerzas_armadas'] or '',
            p['tiene_discapacidad'] or '', p['tipo_discapacidad'] or '', p['created_at']
        ]
        for col_num, val in enumerate(vals, 1):
            ws.cell(row=row_num, column=col_num, value=val)

    aplicar_estilo_excel(ws, postulantes, headers)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"registrados_{datetime.now(TIMEZONE).strftime('%Y%m%d_%H%M%S')}.xlsx")


# ===============================
# CONVOCATORIA — abrir / cerrar
# ===============================
@app.get("/api/convocatoria/estado")
def get_estado_convocatoria():
    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT valor FROM configuracion WHERE clave = 'convocatoria_activa'")
                row = cur.fetchone()
                activa = (row['valor'] == 'true') if row else True
        return jsonify({"ok": True, "activa": activa})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post("/api/convocatoria/estado")
def set_estado_convocatoria():
    err = require_rol("admin")
    if err: return err
    err2 = require_csrf()
    if err2: return err2

    data = request.get_json(silent=True) or {}
    activa = data.get("activa", True)
    valor = 'true' if activa else 'false'

    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO configuracion (clave, valor) VALUES ('convocatoria_activa', %s)
                    ON CONFLICT (clave) DO UPDATE SET valor = EXCLUDED.valor
                """, (valor,))
                conn.commit()
        accion = "Abrió la convocatoria" if activa else "Cerró la convocatoria"
        registrar_log(session.get("usuario", "admin"), accion)
        return jsonify({"ok": True, "activa": activa})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ===============================
# HEARTBEAT — usuarios conectados
# ===============================
@app.post("/api/heartbeat")
def heartbeat():
    err = require_rol("admin", "usuario")
    if err: return err

    username = session.get("usuario")
    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO sesiones_activas (username, ultimo_latido) VALUES (%s, %s)
                    ON CONFLICT (username) DO UPDATE SET ultimo_latido = EXCLUDED.ultimo_latido
                """, (username, now_peru()))
                conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.get("/api/usuarios-activos")
def usuarios_activos():
    err = require_rol("admin")
    if err: return err

    try:
        with PooledConn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT s.username, s.ultimo_latido, u.rol
                    FROM sesiones_activas s
                    JOIN usuarios u ON u.username = s.username
                    WHERE u.activo = 1
                    ORDER BY s.ultimo_latido DESC
                """)
                rows = cur.fetchall()

        ahora = datetime.now(TIMEZONE)
        activos = []
        for r in rows:
            try:
                ultimo = datetime.strptime(r['ultimo_latido'], "%Y-%m-%d %H:%M:%S")
                ultimo = TIMEZONE.localize(ultimo)
                segundos = (ahora - ultimo).total_seconds()
                if segundos <= 90:
                    activos.append({
                        "username": r['username'],
                        "rol": r['rol'],
                        "ultimo_latido": r['ultimo_latido'],
                        "segundos_inactivo": int(segundos)
                    })
            except Exception:
                pass

        return jsonify({"ok": True, "activos": activos})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


if __name__ == "__main__":
    print("=" * 70)
    print("🚀 SISTEMA DE REGISTRO DE POSTULANTES CAS 2026 - MML")
    print("=" * 70)
    print("✅ Connection pooling (psycopg2) — conexiones reutilizadas")
    print("✅ Rate limiting en login — bloqueo tras 5 intentos fallidos")
    print("✅ Protección CSRF — token por sesión en todas las mutaciones")
    print("✅ Timeout de sesión — cierre automático a las 8 horas")
    print("✅ UPDATE atómico en recepción — sin colisiones entre usuarios")
    print("✅ Logout limpia sesiones activas inmediatamente")
    print("💾 Base de datos: PostgreSQL (Azure)")
    print("🌐 Acceso: http://localhost:5000")
    print("=" * 70)
    app.run(debug=False)